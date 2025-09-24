# import streamlit as st
# import os
# import re
# import logging
# import pandas as pd
# from typing import List, Tuple

# import yt_dlp
# from pydub import AudioSegment
# from openai import OpenAI
# from openpyxl import load_workbook

# # ======================================
# # ⚙️ Config
# # ======================================
# st.set_page_config(page_title="YouTube → Transcripts → GPT Summary", layout="wide")
# st.title("YouTube → Transcripts → GPT Summary")

# st.caption("Paste YouTube video links → download audio → transcribe (Whisper) → summarize all transcripts with your custom ChatGPT prompt.")

# # Provide your API key in env var OPENAI_API_KEY or the input below
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
# api_key_input = st.text_input("OpenAI API Key (optional if set in env)", value=OPENAI_API_KEY, type="password")

# # Init OpenAI client lazily
# @st.cache_resource(show_spinner=False)
# def get_client(api_key: str):
#     if not api_key:
#         raise ValueError("Missing OpenAI API Key. Provide via env OPENAI_API_KEY or the input box.")
#     return OpenAI(api_key=api_key)

# # ======================================
# # 🧰 Utilities
# # ======================================
# log_file = "log.txt"
# logging.basicConfig(filename=log_file, level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# def log_message(message: str):
#     logging.info(message)
#     st.write(message)

# INVALID_CHARS = r"\\/:*?\"<>|\n\r\t"

# # Common newline constants to avoid backslashes inside f-string expressions
# NEWLINE = "\n"
# DOUBLE_NL = "\n\n"

# def sanitize_filename(name: str) -> str:
#     """Remove file-system breaking characters only; keep readability."""
#     name = str(name)
#     return re.sub(f"[{INVALID_CHARS}]", " ", name).strip()

# # Timestamp formatter (H:MM:SS or MM:SS)

# def format_time_hms(seconds: float) -> str:
#     hours = int(seconds // 3600)
#     minutes = int((seconds % 3600) // 60)
#     remaining_seconds = int(seconds % 60)
#     if hours:
#         return f"{hours:01}:{minutes:02}:{remaining_seconds:02}"
#     return f"{minutes:02}:{remaining_seconds:02}"

# # Split audio into chunks for Whisper size limits (default 60s)

# def split_audio(file_path: str, chunk_length_ms: int = 60_000) -> Tuple[List[str], List[float]]:
#     audio = AudioSegment.from_file(file_path)
#     chunks = [audio[i:i + chunk_length_ms] for i in range(0, len(audio), chunk_length_ms)]
#     base, _ = os.path.splitext(file_path)

#     chunk_paths, chunk_starts = [], []
#     for i, chunk in enumerate(chunks):
#         out_path = f"{base}_part{i}.mp3"
#         chunk.export(out_path, format="mp3")
#         chunk_paths.append(out_path)
#         chunk_starts.append(i * (chunk_length_ms / 1000.0))
#     return chunk_paths, chunk_starts

# # ======================================
# # 1) INPUT: YouTube links
# # ======================================
# st.subheader("1) Paste YouTube video links (one per line)")
# links_text = st.text_area("YouTube Links", placeholder="https://www.youtube.com/watch?v=...\nhttps://youtu.be/...", height=160)

# save_root = st.text_input("Output folder (will create subfolders: Audio, Transcript, Timestamped)", value=os.path.abspath("./yt_outputs"))

# col_a, col_b = st.columns(2)
# with col_a:
#     download_audio = st.checkbox("Download audio (MP3)", value=True)
# with col_b:
#     download_video = st.checkbox("Download video (MP4)", value=True)

# chunk_minutes = st.number_input("Transcription chunk length (minutes)", min_value=0.1, max_value=10.0, value=0.1, step=0.1)

# # Prepare folders
# AUDIO_DIR = os.path.join(save_root, "Audio")
# TRANSCRIPT_DIR = os.path.join(save_root, "Transcript")
# TIMESTAMP_DIR = os.path.join(save_root, "Timestamped")
# SUMMARY_DIR = os.path.join(save_root, "Summary")
# VIDEO_DIR = os.path.join(save_root, "Video")

# # ======================================
# # 2) Download audio & Transcribe to individual files
# # ======================================
# st.subheader("2) Download transcripts as individual files")

# if st.button("Download & Transcribe"):
#     try:
#         os.makedirs(AUDIO_DIR, exist_ok=True)
#         os.makedirs(TRANSCRIPT_DIR, exist_ok=True)
#         os.makedirs(TIMESTAMP_DIR, exist_ok=True)
#         os.makedirs(SUMMARY_DIR, exist_ok=True)
#         os.makedirs(VIDEO_DIR, exist_ok=True)
#     except Exception as e:
#         st.error(f"Failed to create folders at {save_root}: {e}")
#         st.stop()

#     links = [ln.strip() for ln in links_text.splitlines() if ln.strip()]
#     if not links:
#         st.error("Please paste at least one YouTube link.")
#         st.stop()

#     # Setup yt-dlp
#     def audio_hook(d):
#         if d.get('status') == 'finished':
#             log_message(f"🎵 Audio ready: {d.get('filename')}")

#     def video_hook(d):
#         if d.get('status') == 'finished':
#             log_message(f"📽️ Video ready: {d.get('filename')}")

#     ydl_opts_audio = {
#         'format': 'bestaudio/best',
#         'outtmpl': os.path.join(AUDIO_DIR, '%(title)s.%(ext)s'),
#         'postprocessors': [
#             {'key': 'FFmpegExtractAudio', 'preferredcodec': 'mp3', 'preferredquality': '192'}
#         ],
#         'progress_hooks': [audio_hook],
#         'noplaylist': True,
#     }

#     ydl_opts_video = {
#         'format': 'bestvideo+bestaudio/best',
#         'merge_output_format': 'mp4',
#         'outtmpl': os.path.join(VIDEO_DIR, '%(title)s.%(ext)s'),
#         'progress_hooks': [video_hook],
#         'noplaylist': True,
#     }

#     # Download audio
#     if download_audio:
#         log_message(f"Downloading audio for {len(links)} link(s)...")
#         with yt_dlp.YoutubeDL(ydl_opts_audio) as ydl:
#             ydl.download(links)
#         log_message("✅ Audio download complete.")

#     # Download video
#     if download_video:
#         log_message(f"Downloading video for {len(links)} link(s)...")
#         with yt_dlp.YoutubeDL(ydl_opts_video) as ydl:
#             ydl.download(links)
#         log_message("✅ Video download complete.")

#     # Sanitize any problematic filenames (keep .mp3)
#     if download_audio:
#         for f in os.listdir(AUDIO_DIR):
#             fp = os.path.join(AUDIO_DIR, f)
#             if os.path.isfile(fp):
#                 name, ext = os.path.splitext(f)
#                 new_name = sanitize_filename(name) + ext
#                 new_fp = os.path.join(AUDIO_DIR, new_name)
#                 if new_fp != fp:
#                     try:
#                         os.rename(fp, new_fp)
#                     except Exception:
#                         pass

#     # Sanitize any problematic filenames (keep .mp4)
#     if download_video:
#         for f in os.listdir(VIDEO_DIR):
#             fp = os.path.join(VIDEO_DIR, f)
#             if os.path.isfile(fp):
#                 name, ext = os.path.splitext(f)
#                 new_name = sanitize_filename(name) + ext
#                 new_fp = os.path.join(VIDEO_DIR, new_name)
#                 if new_fp != fp:
#                     try:
#                         os.rename(fp, new_fp)
#                     except Exception:
#                         pass

#     # Transcribe
#     st.write("📜 Starting transcription with Whisper…")
#     client = get_client(api_key_input)

#     records = []  # For an index table

#     for file_name in sorted(os.listdir(AUDIO_DIR)):
#         if not file_name.lower().endswith('.mp3'):
#             continue
#         audio_path = os.path.join(AUDIO_DIR, file_name)
#         size_mb = os.path.getsize(audio_path) / (1024 * 1024)
#         st.write(f"⏳ {file_name} ({size_mb:.2f} MB)")

#         # Split if > 25 MB
#         if size_mb > 25:
#             chunk_paths, chunk_starts = split_audio(audio_path, int(chunk_minutes * 60_000))
#         else:
#             chunk_paths, chunk_starts = [audio_path], [0.0]

#         plain_lines: List[str] = []
#         rows = []  # [start, end, text]

#         for idx, chunk_path in enumerate(chunk_paths):
#             start_offset = chunk_starts[idx]
#             with open(chunk_path, 'rb') as fh:
#                 try:
#                     data = client.audio.transcriptions.create(
#                         model='whisper-1',
#                         file=fh,
#                         response_format='verbose_json'
#                     ).model_dump()
#                 except Exception as e:
#                     st.error(f"Transcription failed for {os.path.basename(chunk_path)}: {e}")
#                     continue

#             for seg in data.get('segments', []) or []:
#                 s = float(seg.get('start', 0.0)) + start_offset
#                 e = float(seg.get('end', 0.0)) + start_offset
#                 text = str(seg.get('text', '')).strip()
#                 rows.append([format_time_hms(s), format_time_hms(e), text])
#                 plain_lines.append(text)

#             # Remove temp chunks
#             if chunk_path != audio_path:
#                 try:
#                     os.remove(chunk_path)
#                 except Exception:
#                     pass

#         base = os.path.splitext(file_name)[0]
#         txt_path = os.path.join(TRANSCRIPT_DIR, f"{base}_transcript.txt")
#         xlsx_path = os.path.join(TIMESTAMP_DIR, f"{base}_timestamp.xlsx")

#         # Save plain transcript
#         with open(txt_path, 'w', encoding='utf-8') as f:
#             f.write("\n".join(plain_lines))

#         # Save timestamped transcript
#         df = pd.DataFrame(rows, columns=["Start", "End", "Text"]) if rows else pd.DataFrame(columns=["Start", "End", "Text"])
#         df.to_excel(xlsx_path, index=False)
#         try:
#             wb = load_workbook(xlsx_path)
#             ws = wb.active
#             ws.column_dimensions['A'].width = 12
#             ws.column_dimensions['B'].width = 12
#             ws.column_dimensions['C'].width = 120
#             wb.save(xlsx_path)
#         except Exception:
#             pass

#         log_message(f"📄 Saved: {txt_path}")
#         log_message(f"📄 Saved: {xlsx_path}")
#         records.append({"File": file_name, "Transcript": txt_path, "Timestamped": xlsx_path, "SizeMB": f"{size_mb:.2f}"})

#     if records:
#         st.success("All transcripts saved.")
#         st.dataframe(pd.DataFrame(records))
#     else:
#         st.warning("No transcripts were generated.")

# # ======================================
# # 3) Summarize all transcripts with a custom prompt
# # ======================================
# st.subheader("3) Summarize using a custom ChatGPT prompt")
# user_prompt = st.text_area(
#     "Your summarization prompt (the transcripts will be appended after this prompt)",
#     value="""以下の複数の文字起こしを読み、要点を日本語で箇条書き要約してください。
# - 事実と意見を分けて記載
# - 時系列の整理
# - 再利用できるキーメッセージ5つ
# - 改善アクション（チャンネル運営の観点）
# """,
#     height=180,
# )

# opt1, opt2, opt3, opt4 = st.columns(4)
# with opt1:
#     synthesize_level = st.selectbox(
#         "Synthesis strategy",
#         ["Auto (per-file summary → final synthesis)", "Single pass (concatenate all)"],
#         index=0,
#     )
# with opt2:
#     max_chars_per_call = st.number_input(
#         "Max transcript chars per model call",
#         min_value=2000,
#         max_value=30000,
#         value=12000,
#         step=1000,
#     )
# with opt3:
#     gen_per_file = st.checkbox("Per-transcript summaries", value=True)
#     gen_aggregated = st.checkbox("Aggregated summary", value=True)
# with opt4:
#     export_table_csv = st.checkbox("Export per-file table (CSV)", value=True)
#     export_table_xlsx = st.checkbox("Export per-file table (Excel)", value=False)

# if st.button("Run Summary"):
#     if not (gen_per_file or gen_aggregated):
#         st.warning("Please select at least one output: per-transcript and/or aggregated.")
#         st.stop()

#     client = get_client(api_key_input)

#     if not os.path.isdir(TRANSCRIPT_DIR):
#         st.error(f"Transcript folder not found: {TRANSCRIPT_DIR}. Run step 2 first.")
#         st.stop()

#     txt_files = [os.path.join(TRANSCRIPT_DIR, f) for f in os.listdir(TRANSCRIPT_DIR) if f.lower().endswith('.txt')]
#     if not txt_files:
#         st.error("No transcript .txt files found. Run step 2 first.")
#         st.stop()

#     st.write(f"Found {len(txt_files)} transcript(s).")

#     def chat(prompt: str) -> str:
#         try:
#             resp = client.chat.completions.create(
#                 model="gpt-4o",
#                 messages=[
#                     {"role": "system", "content": "You are a concise, expert analyst."},
#                     {"role": "user", "content": prompt}
#                 ]
#             )
#             return resp.choices[0].message.content
#         except Exception as e:
#             return f"[ERROR] {e}"

#     def summarize_one(title: str, content: str) -> str:
#         """Summarize one transcript safely (no backslashes inside f-string expressions)."""
#         if len(content) <= max_chars_per_call:
#             prompt = f"""{user_prompt}

# 【対象トランスクリプト: {title}】

# {content}
# """
#             return chat(prompt)
#         chunks = [content[i:i+max_chars_per_call] for i in range(0, len(content), max_chars_per_call)]
#         part_summaries = []
#         for i, ch in enumerate(chunks, 1):
#             prompt = f"""{user_prompt}

# 【対象トランスクリプト: {title} / Part {i}/{len(chunks)}】

# {ch}
# """
#             part_summaries.append(chat(prompt))
#         sep = DOUBLE_NL
#         joined_parts = sep.join(part_summaries)
#         synth_prompt = f"""以下は同じ動画文字起こしの部分要約です。重複を避け、網羅的に1つの要約に統合してください。

# {joined_parts}
# """
#         return chat(synth_prompt)

#     per_file_records = []
#     intermediate_summaries = []

#     if synthesize_level.startswith("Auto"):
#         for path in sorted(txt_files):
#             title = os.path.splitext(os.path.basename(path))[0]
#             with open(path, "r", encoding="utf-8") as f:
#                 content = f.read()
#             st.write(f"Summarizing: {title}")
#             summary_text = summarize_one(title, content)
#             intermediate_summaries.append(f"""# Summary: {title}

# {summary_text}

# """)
#             if gen_per_file:
#                 out_path = os.path.join(SUMMARY_DIR, f"{title}_summary.md")
#                 with open(out_path, "w", encoding="utf-8") as w:
#                     w.write(summary_text)
#                 per_file_records.append({"Title": title, "Summary": summary_text, "Summary File": out_path, "Length": len(summary_text)})

#         final_summary = None
#         if gen_aggregated:
#             sep = DOUBLE_NL
#             joined_intermediate = sep.join(intermediate_summaries)
#             final_prompt = f"""以下は複数動画の要約です。全体としてのハイライト・矛盾点・学び・推奨アクションを日本語でまとめてください。

# {joined_intermediate}
# """
#             final_summary = chat(final_prompt)

#     else:
#         if gen_per_file:
#             for path in sorted(txt_files):
#                 title = os.path.splitext(os.path.basename(path))[0]
#                 with open(path, "r", encoding="utf-8") as f:
#                     content = f.read()
#                 st.write(f"Summarizing (single pass per file): {title}")
#                 summary_text = summarize_one(title, content)
#                 intermediate_summaries.append(f"""# Summary: {title}

# {summary_text}

# """)
#                 out_path = os.path.join(SUMMARY_DIR, f"{title}_summary.md")
#                 with open(out_path, "w", encoding="utf-8") as w:
#                     w.write(summary_text)
#                 per_file_records.append({"Title": title, "Summary": summary_text, "Summary File": out_path, "Length": len(summary_text)})

#         final_summary = None
#         if gen_aggregated:
#             if intermediate_summaries:
#                 sep = DOUBLE_NL
#                 joined_intermediate = sep.join(intermediate_summaries)
#                 final_prompt = f"""以下は複数動画の要約です。全体としてのハイライト・矛盾点・学び・推奨アクションを日本語でまとめてください。

# {joined_intermediate}
# """
#                 final_summary = chat(final_prompt)
#             else:
#                 corpus = []
#                 for path in sorted(txt_files):
#                     with open(path, "r", encoding="utf-8") as f:
#                         text = f.read()
#                     corpus.append(f"""# {os.path.basename(path)}

# {text}
# """)
#                 sep = DOUBLE_NL
#                 big_text = sep.join(corpus)
#                 final_summary = chat(f"""{user_prompt}

# {big_text}
# """)

#     if per_file_records:
#         st.success("✅ Per-transcript summaries generated.")
#         df_idx = pd.DataFrame(per_file_records)
#         st.dataframe(df_idx)
#         for row in per_file_records[:50]:
#             try:
#                 with open(row["Summary File"], "r", encoding="utf-8") as fh:
#                     st.download_button(
#                         label=f"Download {os.path.basename(row['Summary File'])}",
#                         data=fh.read(),
#                         file_name=os.path.basename(row["Summary File"]),
#                         mime="text/markdown",
#                     )
#             except Exception:
#                 pass

#     if gen_aggregated and (final_summary is not None):
#         out_md = os.path.join(SUMMARY_DIR, "summary.md")
#         with open(out_md, "w", encoding="utf-8") as f:
#             f.write(final_summary)
#         st.success("✅ Aggregated summary complete.")
#         st.download_button("Download summary.md", data=final_summary, file_name="summary.md", mime="text/markdown")
#         st.markdown("---")
#         st.markdown("### Final Summary")
#         st.markdown(final_summary)

#     if not per_file_records and (final_summary is None):
#         st.warning("No summaries were produced. Check your options and try again.")

# # ======================================
# # Optional lightweight sanity tests
# # Run only if RUN_APP_TESTS=1 is set in the environment
# # ======================================

# def _run_sanity_tests():
#     # Ensure newline joins behave as expected and no syntax pitfalls remain
#     assert NEWLINE == "\n"
#     assert DOUBLE_NL == "\n\n"
#     parts = ["alpha", "beta", "gamma"]
#     joined = DOUBLE_NL.join(parts)
#     assert joined == "alpha\n\nbeta\n\ngamma"
#     inter = ["# Summary: A", "# Summary: B"]
#     joined2 = DOUBLE_NL.join(inter)
#     assert "# Summary: A" in joined2 and "# Summary: B" in joined2

# if os.getenv("RUN_APP_TESTS") == "1":
#     try:
#         _run_sanity_tests()
#         print("Sanity tests passed.")
#     except AssertionError as e:
#         print(f"Sanity tests failed: {e}")
import streamlit as st
import os
import re
import logging
import pandas as pd
from typing import List, Tuple

import yt_dlp
from pydub import AudioSegment
from openai import OpenAI
from openpyxl import load_workbook

# ======================================
# ⚙️ Config
# ======================================
st.set_page_config(page_title="YouTube → Transcripts → GPT Summary", layout="wide")
st.title("YouTube → Transcripts → GPT Summary")

st.caption("Paste YouTube video links → download audio/video → transcribe (Whisper) → summarize transcripts with your custom ChatGPT prompt.")

# Provide your API key in env var OPENAI_API_KEY or the input below
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
api_key_input = st.text_input("OpenAI API Key (optional if set in env)", value=OPENAI_API_KEY, type="password")

# Init OpenAI client lazily
@st.cache_resource(show_spinner=False)
def get_client(api_key: str):
    if not api_key:
        raise ValueError("Missing OpenAI API Key. Provide via env OPENAI_API_KEY or the input box.")
    return OpenAI(api_key=api_key)

# ======================================
# 🧰 Utilities
# ======================================
log_file = "log.txt"
logging.basicConfig(filename=log_file, level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

def log_message(message: str):
    logging.info(message)
    st.write(message)

# Invalid filename characters (Windows-safe) — use single quotes to allow an unescaped double-quote
# Include backslash, forward slash, colon, asterisk, question mark, double-quote, angle brackets, pipe, newline, CR, tab
INVALID_CHARS = r'\\/:*?"<>|\n\r\t'

# Common newline constants to avoid backslashes inside f-string expressions
NEWLINE = "\n"
DOUBLE_NL = "\n\n"

def sanitize_filename(name: str) -> str:
    """Remove file-system breaking characters only; keep readability."""
    name = str(name)
    return re.sub(f"[{INVALID_CHARS}]", " ", name).strip()

def normalize_root(path_str: str) -> str:
    """Normalize user-entered folder paths (Windows-friendly).
    - Strips wrapping quotes
    - Expands env vars and user (~)
    - Normalizes separators and returns absolute path
    """
    if not path_str:
        path_str = "./yt_outputs"
    p = path_str.strip()
    if (p.startswith('"') and p.endswith('"')) or (p.startswith("'") and p.endswith("'")):
        p = p[1:-1]
    p = os.path.expandvars(os.path.expanduser(p))
    p = os.path.abspath(os.path.normpath(p))
    return p

def ensure_dirs(*paths: str) -> None:
    for p in paths:
        try:
            os.makedirs(p, exist_ok=True)
        except Exception:
            pass

# Timestamp formatter (H:MM:SS or MM:SS)

def format_time_hms(seconds: float) -> str:
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    remaining_seconds = int(seconds % 60)
    if hours:
        return f"{hours:01}:{minutes:02}:{remaining_seconds:02}"
    return f"{minutes:02}:{remaining_seconds:02}"

# Split audio into chunks for Whisper size limits (default 60s)

def split_audio(file_path: str, chunk_length_ms: int = 60_000) -> Tuple[List[str], List[float]]:
    audio = AudioSegment.from_file(file_path)
    chunks = [audio[i:i + chunk_length_ms] for i in range(0, len(audio), chunk_length_ms)]
    base, _ = os.path.splitext(file_path)

    chunk_paths, chunk_starts = [], []
    for i, chunk in enumerate(chunks):
        out_path = f"{base}_part{i}.mp3"
        chunk.export(out_path, format="mp3")
        chunk_paths.append(out_path)
        chunk_starts.append(i * (chunk_length_ms / 1000.0))
    return chunk_paths, chunk_starts

# ======================================
# 1) INPUT: YouTube links
# ======================================
st.subheader("1) Paste YouTube video links (one per line)")
links_text = st.text_area(
    "YouTube Links",
    placeholder="https://www.youtube.com/watch?v=...\nhttps://youtu.be/...",
    height=160,
)

save_root = st.text_input(
    "Output folder (will create subfolders: Audio, Transcript, Timestamped, Summary, Video)",
    value=os.path.abspath("./yt_outputs"),
)
root_dir = normalize_root(save_root)
st.caption(f"Resolved output root: {root_dir}")

col_a, col_b = st.columns(2)
with col_a:
    download_audio = st.checkbox("Download audio (MP3)", value=True)
with col_b:
    download_video = st.checkbox("Download video (MP4)", value=True)

chunk_minutes = st.number_input("Transcription chunk length (minutes)", min_value=0.1, max_value=10.0, value=0.1, step=0.1)

# Prepare folders
AUDIO_DIR = os.path.join(root_dir, "Audio")
TRANSCRIPT_DIR = os.path.join(root_dir, "Transcript")
TIMESTAMP_DIR = os.path.join(root_dir, "Timestamped")
SUMMARY_DIR = os.path.join(root_dir, "Summary")
VIDEO_DIR = os.path.join(root_dir, "Video")

# ======================================
# 2) Download audio & Transcribe to individual files
# ======================================
st.subheader("2) Download transcripts as individual files")

if st.button("Download & Transcribe"):
    try:
        ensure_dirs(AUDIO_DIR, TRANSCRIPT_DIR, TIMESTAMP_DIR, SUMMARY_DIR, VIDEO_DIR)
        st.success(f"Output folders ready under: {root_dir}")
    except Exception as e:
        st.error(f"Failed to create folders at {root_dir}: {e}")
        st.stop()

    links = [ln.strip() for ln in links_text.splitlines() if ln.strip()]
    if not links:
        st.error("Please paste at least one YouTube link.")
        st.stop()

    # Setup yt-dlp
    def audio_hook(d):
        if d.get('status') == 'finished':
            log_message(f"🎵 Audio ready: {d.get('filename')}")

    def video_hook(d):
        if d.get('status') == 'finished':
            log_message(f"📽️ Video ready: {d.get('filename')}")

    ydl_opts_audio = {
        'format': 'bestaudio/best',
        'outtmpl': os.path.join(AUDIO_DIR, '%(title)s-%(id)s.%(ext)s'),
        'postprocessors': [
            {'key': 'FFmpegExtractAudio', 'preferredcodec': 'mp3', 'preferredquality': '192'}
        ],
        'progress_hooks': [audio_hook],
        'noplaylist': True,
        'restrictfilenames': True,
        'windowsfilenames': True,
    }

    ydl_opts_video = {
        'format': 'bestvideo*+bestaudio/best',
        'merge_output_format': 'mp4',
        'outtmpl': os.path.join(VIDEO_DIR, '%(title)s-%(id)s.%(ext)s'),
        'progress_hooks': [video_hook],
        'noplaylist': True,
        'restrictfilenames': True,
        'windowsfilenames': True,
    }

    # Download audio
    if download_audio:
        log_message(f"Downloading audio for {len(links)} link(s)... → {AUDIO_DIR}")
        with yt_dlp.YoutubeDL(ydl_opts_audio) as ydl:
            ydl.download(links)
        log_message("✅ Audio download complete.")

    # Download video
    if download_video:
        log_message(f"Downloading video for {len(links)} link(s)... → {VIDEO_DIR}")
        with yt_dlp.YoutubeDL(ydl_opts_video) as ydl:
            ydl.download(links)
        log_message("✅ Video download complete.")

    # Sanitize any problematic filenames (keep .mp3)
    if download_audio:
        for f in os.listdir(AUDIO_DIR):
            fp = os.path.join(AUDIO_DIR, f)
            if os.path.isfile(fp):
                name, ext = os.path.splitext(f)
                new_name = sanitize_filename(name) + ext
                new_fp = os.path.join(AUDIO_DIR, new_name)
                if new_fp != fp:
                    try:
                        os.rename(fp, new_fp)
                    except Exception:
                        pass

    # Sanitize any problematic filenames (keep .mp4)
    if download_video:
        for f in os.listdir(VIDEO_DIR):
            fp = os.path.join(VIDEO_DIR, f)
            if os.path.isfile(fp):
                name, ext = os.path.splitext(f)
                new_name = sanitize_filename(name) + ext
                new_fp = os.path.join(VIDEO_DIR, new_name)
                if new_fp != fp:
                    try:
                        os.rename(fp, new_fp)
                    except Exception:
                        pass

    # Transcribe
    st.write("📜 Starting transcription with Whisper…")
    client = get_client(api_key_input)

    records = []  # For an index table

    for file_name in sorted(os.listdir(AUDIO_DIR)):
        if not file_name.lower().endswith('.mp3'):
            continue
        audio_path = os.path.join(AUDIO_DIR, file_name)
        size_mb = os.path.getsize(audio_path) / (1024 * 1024)
        st.write(f"⏳ {file_name} ({size_mb:.2f} MB)")

        # Split if > 25 MB
        if size_mb > 25:
            chunk_paths, chunk_starts = split_audio(audio_path, int(chunk_minutes * 60_000))
        else:
            chunk_paths, chunk_starts = [audio_path], [0.0]

        plain_lines: List[str] = []
        rows = []  # [start, end, text]

        for idx, chunk_path in enumerate(chunk_paths):
            start_offset = chunk_starts[idx]
            with open(chunk_path, 'rb') as fh:
                try:
                    data = client.audio.transcriptions.create(
                        model='whisper-1',
                        file=fh,
                        response_format='verbose_json'
                    ).model_dump()
                except Exception as e:
                    st.error(f"Transcription failed for {os.path.basename(chunk_path)}: {e}")
                    continue

            for seg in data.get('segments', []) or []:
                s = float(seg.get('start', 0.0)) + start_offset
                e = float(seg.get('end', 0.0)) + start_offset
                text = str(seg.get('text', '')).strip()
                rows.append([format_time_hms(s), format_time_hms(e), text])
                plain_lines.append(text)

            # Remove temp chunks
            if chunk_path != audio_path:
                try:
                    os.remove(chunk_path)
                except Exception:
                    pass

        base = os.path.splitext(file_name)[0]
        txt_path = os.path.join(TRANSCRIPT_DIR, f"{base}_transcript.txt")
        xlsx_path = os.path.join(TIMESTAMP_DIR, f"{base}_timestamp.xlsx")

        # Save plain transcript
        ensure_dirs(TRANSCRIPT_DIR)
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write(NEWLINE.join(plain_lines))

        # Save timestamped transcript
        ensure_dirs(TIMESTAMP_DIR)
        df = pd.DataFrame(rows, columns=["Start", "End", "Text"]) if rows else pd.DataFrame(columns=["Start", "End", "Text"])
        df.to_excel(xlsx_path, index=False)
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 120
            wb.save(xlsx_path)
        except Exception:
            pass

        log_message(f"📄 Saved: {txt_path}")
        log_message(f"📄 Saved: {xlsx_path}")
        records.append({"File": file_name, "Transcript": txt_path, "Timestamped": xlsx_path, "SizeMB": f"{size_mb:.2f}"})

    if records:
        st.success("All transcripts saved.")
        st.dataframe(pd.DataFrame(records))
    else:
        st.warning("No transcripts were generated.")

# ======================================
# 3) Summarize all transcripts with a custom prompt
# ======================================
st.subheader("3) Summarize using a custom ChatGPT prompt")
user_prompt = st.text_area(
    "Your summarization prompt (the transcripts will be appended after this prompt)",
    value="""以下の複数の文字起こしを読み、要点を日本語で箇条書き要約してください。
- 事実と意見を分けて記載
- 時系列の整理
- 再利用できるキーメッセージ5つ
- 改善アクション（チャンネル運営の観点）
""",
    height=180,
)

opt1, opt2, opt3, opt4 = st.columns(4)
with opt1:
    synthesize_level = st.selectbox(
        "Synthesis strategy",
        ["Auto (per-file summary → final synthesis)", "Single pass (concatenate all)"],
        index=0,
    )
with opt2:
    max_chars_per_call = st.number_input(
        "Max transcript chars per model call",
        min_value=2000,
        max_value=30000,
        value=12000,
        step=1000,
    )
with opt3:
    gen_per_file = st.checkbox("Per-transcript summaries", value=True)
    gen_aggregated = st.checkbox("Aggregated summary", value=True)
with opt4:
    export_table_csv = st.checkbox("Export per-file table (CSV)", value=True)
    export_table_xlsx = st.checkbox("Export per-file table (Excel)", value=False)

# Ensure Summary dir exists even if user didn't run Step 2
ensure_dirs(SUMMARY_DIR)

if st.button("Run Summary"):
    if not (gen_per_file or gen_aggregated):
        st.warning("Please select at least one output: per-transcript and/or aggregated.")
        st.stop()

    client = get_client(api_key_input)

    if not os.path.isdir(TRANSCRIPT_DIR):
        st.error(f"Transcript folder not found: {TRANSCRIPT_DIR}. Run step 2 first.")
        st.stop()

    txt_files = [os.path.join(TRANSCRIPT_DIR, f) for f in os.listdir(TRANSCRIPT_DIR) if f.lower().endswith('.txt')]
    if not txt_files:
        st.error("No transcript .txt files found. Run step 2 first.")
        st.stop()

    st.write(f"Found {len(txt_files)} transcript(s).")

    def chat(prompt: str) -> str:
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a concise, expert analyst."},
                    {"role": "user", "content": prompt}
                ]
            )
            return resp.choices[0].message.content
        except Exception as e:
            return f"[ERROR] {e}"

    def summarize_one(title: str, content: str) -> str:
        """Summarize one transcript safely (no backslashes inside f-string expressions)."""
        if len(content) <= max_chars_per_call:
            prompt = f"""{user_prompt}

【対象トランスクリプト: {title}】

{content}
"""
            return chat(prompt)
        chunks = [content[i:i+max_chars_per_call] for i in range(0, len(content), max_chars_per_call)]
        part_summaries = []
        for i, ch in enumerate(chunks, 1):
            prompt = f"""{user_prompt}

【対象トランスクリプト: {title} / Part {i}/{len(chunks)}】

{ch}
"""
            part_summaries.append(chat(prompt))
        joined_parts = DOUBLE_NL.join(part_summaries)
        synth_prompt = f"""以下は同じ動画文字起こしの部分要約です。重複を避け、網羅的に1つの要約に統合してください。

{joined_parts}
"""
        return chat(synth_prompt)

    per_file_records = []
    intermediate_summaries = []

    if synthesize_level.startswith("Auto"):
        for path in sorted(txt_files):
            title = os.path.splitext(os.path.basename(path))[0]
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            st.write(f"Summarizing: {title}")
            summary_text = summarize_one(title, content)
            intermediate_summaries.append(f"""# Summary: {title}

{summary_text}

""")
            if gen_per_file:
                out_path = os.path.join(SUMMARY_DIR, f"{title}_summary.md")
                with open(out_path, "w", encoding="utf-8") as w:
                    w.write(summary_text)
                per_file_records.append({"Title": title, "Summary": summary_text, "Summary File": out_path, "Length": len(summary_text)})

        final_summary = None
        if gen_aggregated:
            joined_intermediate = DOUBLE_NL.join(intermediate_summaries)
            final_prompt = f"""以下は複数動画の要約です。全体としてのハイライト・矛盾点・学び・推奨アクションを日本語でまとめてください。

{joined_intermediate}
"""
            final_summary = chat(final_prompt)

    else:
        if gen_per_file:
            for path in sorted(txt_files):
                title = os.path.splitext(os.path.basename(path))[0]
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()
                st.write(f"Summarizing (single pass per file): {title}")
                summary_text = summarize_one(title, content)
                intermediate_summaries.append(f"""# Summary: {title}

{summary_text}

""")
                out_path = os.path.join(SUMMARY_DIR, f"{title}_summary.md")
                with open(out_path, "w", encoding="utf-8") as w:
                    w.write(summary_text)
                per_file_records.append({"Title": title, "Summary": summary_text, "Summary File": out_path, "Length": len(summary_text)})

        final_summary = None
        if gen_aggregated:
            if intermediate_summaries:
                joined_intermediate = DOUBLE_NL.join(intermediate_summaries)
                final_prompt = f"""以下は複数動画の要約です。全体としてのハイライト・矛盾点・学び・推奨アクションを日本語でまとめてください。

{joined_intermediate}
"""
                final_summary = chat(final_prompt)
            else:
                corpus = []
                for path in sorted(txt_files):
                    with open(path, "r", encoding="utf-8") as f:
                        text = f.read()
                    corpus.append(f"""# {os.path.basename(path)}

{text}
""")
                big_text = DOUBLE_NL.join(corpus)
                final_summary = chat(f"""{user_prompt}

{big_text}
""")

    if per_file_records:
        st.success("✅ Per-transcript summaries generated.")
        df_idx = pd.DataFrame(per_file_records)
        st.dataframe(df_idx)
        for row in per_file_records[:50]:
            try:
                with open(row["Summary File"], "r", encoding="utf-8") as fh:
                    st.download_button(
                        label=f"Download {os.path.basename(row['Summary File'])}",
                        data=fh.read(),
                        file_name=os.path.basename(row["Summary File"]),
                        mime="text/markdown",
                    )
            except Exception:
                pass

    if gen_aggregated and (final_summary is not None):
        out_md = os.path.join(SUMMARY_DIR, "summary.md")
        with open(out_md, "w", encoding="utf-8") as f:
            f.write(final_summary)
        st.success("✅ Aggregated summary complete.")
        st.download_button("Download summary.md", data=final_summary, file_name="summary.md", mime="text/markdown")
        st.markdown("---")
        st.markdown("### Final Summary")
        st.markdown(final_summary)

    if not per_file_records and (final_summary is None):
        st.warning("No summaries were produced. Check your options and try again.")

# ======================================
# Optional lightweight sanity tests
# Run only if RUN_APP_TESTS=1 is set in the environment
# ======================================

def _run_sanity_tests():
    # Ensure newline joins behave as expected and no syntax pitfalls remain
    assert NEWLINE == "\n"
    assert DOUBLE_NL == "\n\n"
    parts = ["alpha", "beta", "gamma"]
    joined = DOUBLE_NL.join(parts)
    assert joined == "alpha\n\nbeta\n\ngamma"
    inter = ["# Summary: A", "# Summary: B"]
    joined2 = DOUBLE_NL.join(inter)
    assert "# Summary: A" in joined2 and "# Summary: B" in joined2

    # Additional tests: sanitize_filename and normalize_root
    dirty = "a/b:c*?\"<>|\n\r\tb"
    cleaned = sanitize_filename(dirty)
    # No invalid characters remain
    assert not re.search(f"[{INVALID_CHARS}]", cleaned)

    # normalize_root strips quotes and returns absolute path
    nr = normalize_root('"./yt_outputs"')
    assert not nr.startswith('"') and os.path.isabs(nr)

if os.getenv("RUN_APP_TESTS") == "1":
    try:
        _run_sanity_tests()
        print("Sanity tests passed.")
    except AssertionError as e:
        print(f"Sanity tests failed: {e}")
